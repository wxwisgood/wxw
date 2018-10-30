# -*- coding: utf-8 -*-

import os
import random
import xml.etree.ElementTree as ET
import time
import openpyxl
import numpy as np

def creat_train_data_list():
    '''生成列表, 将所有数据按文件夹放入Data路径下'''
    folder_input='Data'
    folder_names=os.listdir(folder_input) #寻找所有的文件夹
    folder_names_real=[]
    for folder_name in folder_names:
        if(os.path.isdir(os.path.join(folder_input,folder_name))):
            folder_names_real.append(os.path.join(folder_input,folder_name))

    ftrainval = open(folder_input+'/trainval_list.txt', 'a+')
    ftest = open(folder_input+'/test_list.txt', 'a+')
    ftrain = open(folder_input+'/train_list.txt', 'a+')
    fval = open(folder_input+'/val_list.txt', 'a+')

    total_xml= set()   #遍历所有的文件夹, 按照所有的.xml 来将所有的数据分类
    count_dict = {}  #计数
    for folder_name in folder_names_real:
        print('读取文件夹:%s'%folder_name)
        log_file.write('读取文件夹:%s \n'%folder_name)

        filenames=os.listdir(folder_name)
        #先判断是否一个图片对应一个标注文件
        if(len(filenames)%2==1):
            print('文件数目有误:%s'%folder_name)
            log_file.write('文件数目有误:%s \n'%folder_name)
            continue

        total_xml1 = set()
        total_xml2 = set()
        for item in filenames:
            if item.endswith('.jpg'):
                total_xml1.add(os.path.join(folder_name,item[:-4]))
            elif item.endswith('.xml'):
                total_xml2.add(os.path.join(folder_name,item[:-4]))
            else:
                print('文件夹 %s 中文件格式有误,请放入.jpg或.xml文件类型'%folder_name)
                log_file.write('文件夹 %s 中文件格式有误,请放入.jpg或.xml文件类型 \n'%folder_name)
                break
        if total_xml1 == total_xml2:
            total_xml |= total_xml1
        else:
            print("请检查文件夹 %s 中.jpg和.xml文件是不是一一对应,跳过读写该文件夹"%folder_name)
            log_file.write("请检查文件夹 %s 中.jpg和.xml文件是不是一一对应,跳过读写该文件夹 \n"%folder_name)
            continue

        print("将文件夹 {} 分成训练/验证/测试集".format(folder_name))
        log_file.write("将文件夹 {} 分成训练/验证/测试集 \n".format(folder_name))

        trainval_percent = 0.2#20% 用来验证 80% 用来训练
        test_val_percent = 0.4 #30%用来测试 70%用来验证

        #随机打乱次序
        total_xml1 = list(total_xml1)
        random.shuffle(total_xml1)

        num = len(total_xml1)
        list_num = range(num)
        tv = max(int(num * trainval_percent) , 13)
        tr = int(tv * test_val_percent)
        trainval = random.sample(list_num, tv)
        train = random.sample(trainval, tr)

        count_train = 0
        count_val = 0
        count_test = 0

        for i in list_num:
            name = total_xml1[i] + '\n'
            if i in trainval:
                ftrainval.write(name)
                if i in train:
                    ftest.write(name)
                    count_test += 1
                else:
                    fval.write(name)
                    count_val += 1
            else:
                ftrain.write(name)
                count_train +=1
        count_dict[folder_name] = [num,count_train,count_val,count_test]

    ftrainval.close()
    ftrain.close()
    fval.close()
    ftest.close()

    total_sample = 0
    train_sample = 0
    val_sample = 0
    test_sample = 0
    for key,value in count_dict.items():
        print('文件夹:{:25s} 文件共:{:4d} ,分为 train:{:4d} ,val:{:4d} , test:{:4d} '.format(key,value[0],value[1],value[2],value[3]))
        log_file.write('文件夹:{:25s} 文件共:{:4d} ,分为 train:{:4d} ,val:{:4d} , test:{:4d} \n'.format(key,value[0],value[1],value[2],value[3]))
        total_sample += value[0]
        train_sample += value[1]
        val_sample += value[2]
        test_sample += value[3]
    print('样本文件中共:{:6d} ,分为train:{:4d} ,val:{:4d} , test:{:4d} '.format(total_sample,train_sample,val_sample,test_sample))
    log_file.write('样本中文件共:{:6d} ,分为train:{:4d} ,val:{:4d} , test:{:4d} \n'.format(total_sample,train_sample,val_sample,test_sample))

    total_xml = list(total_xml)
    random.shuffle(total_xml)
    # 生成全部样本list
    with open(folder_input+'/all_list.txt', 'w') as f_all:
        for item in total_xml:
            f_all.write(item+'\n')

#生成训练数据
def creat_train_data():
    sets = ['train_list','val_list','test_list']
    classes = open('Data/object_classes'+str(times)+'.txt').read().strip().split()
    print('本次训练放入的类别:{}'.format(classes))
    log_file.write('本次训练放入的类别:{} \n'.format(classes))

    def convert_annotation(image_id, list_file, all_train):
        in_file = open('%s.xml'%image_id,encoding = 'utf8')
        tree=ET.parse(in_file)
        root = tree.getroot()

        for obj in root.iter('object'):
            difficult = obj.find('difficult').text
            cls = obj.find('name').text
            if cls not in classes or int(difficult)==1:
                continue
            cls_id = classes.index(cls)
            xmlbox = obj.find('bndbox')
            b = (int(xmlbox.find('xmin').text), int(xmlbox.find('ymin').text), int(xmlbox.find('xmax').text), int(xmlbox.find('ymax').text))
            list_file.write(" " + ",".join([str(a) for a in b]) + ',' + str(cls_id))
            all_train.write(" " + ",".join([str(a) for a in b]) + ',' + str(cls_id))

    all_train = open('Data/all_train_data.txt','w')
    wrong_imgeids = open('Data/%s.txt'%('wrong_sample'+str(times))).read().strip().split('\n')
    for name in sets:
        #读取数据
        image_ids = open('Data/%s.txt'%name).read().strip().split('\n')
        #print(image_ids[0])
        #print(image_ids)
        name_output=name.split('_')[0]
        list_file = open('Data/%s.txt'%name_output, 'w')
        for image_id in image_ids:
            if image_id in wrong_imgeids:
                continue
            else:
                list_file.write('%s.jpg'%image_id)
                all_train.write('%s.jpg'%image_id)
                #print(image_id)
                convert_annotation(image_id, list_file, all_train)
                list_file.write('\n')
                all_train.write('\n')
        list_file.close()
    all_train.close()

def get_classes(classes_path):
    '''加载已有类别真值表,
       返回:列表class_names, 字典class_to_id, 字典class_to_value'''
    class_to_value = {}  #类别和对应的真值表(1/0/-1)
    with open(classes_path) as f:
        class_names = f.readlines()
    for items in class_names:
        item = items.strip().split()
        class_to_value[item[0]] = item[1]
    class_names = [c.strip().split()[0] for c in class_names]
    return class_names , class_to_value

def output_class_list(num=5 , class_not_train=[]):

    def class_count(path_all_list,class_name_total):
        wrong_count = 0
        dict_obeject_classes = {}
        image_ids = open(path_all_list).read().strip().split('\n')
        sum_sample = len(image_ids)
        for image_id in image_ids: 
            in_file = open('%s.xml'%image_id,encoding = 'utf8')
            tree=ET.parse(in_file)
            root = tree.getroot()
            for obj in root.iter('object'):
                difficult = obj.find('difficult').text
                cls = obj.find('name').text
                if (cls in class_name_total) and (cls not in dict_obeject_classes):
                    dict_obeject_classes[cls] = 1

                elif cls in dict_obeject_classes:
                    dict_obeject_classes[cls] += 1					

                elif cls not in class_name_total or int(difficult)==1:
                    wrong_sample.write(image_id+'\n')
                    print('标错样本:%s'%image_id)
                    log_file.write('标签标错样本:%s \n'%image_id)
                    wrong_count += 1
                    continue

        print('标注正确样本:{} 个,标签标错样本共:{} 个 \n'.format(sum_sample-wrong_count, wrong_count))
        log_file.write('标注正确样本:{} 个,标签标错样本共:{} 个 \n'.format(sum_sample-wrong_count, wrong_count))
        return dict_obeject_classes

    def object_class_count(path_all_list,class_name_total):
        obeject_class_count = {}
        image_ids = open(path_all_list).read().strip().split('\n')
        for image_id in image_ids: 
            in_file = open('%s.xml'%image_id,encoding = 'utf8')
            tree=ET.parse(in_file)
            root = tree.getroot()
            for obj in root.iter('object'):
                difficult = obj.find('difficult').text
                cls = obj.find('name').text
                if (cls in class_name_total) and (cls not in obeject_class_count):
                    obeject_class_count[cls] = 1
                elif cls in obeject_class_count:
                    obeject_class_count[cls] += 1
                else:
                    continue
        return obeject_class_count

    path_class_name_all = 'object_classes_all.txt' #所有类别名字文件

    #准备数据及路径
    class_names, class_to_value = get_classes(path_class_name_all)

    #输入文件路径
    path_all_list =  'Data/all_list.txt'  #所有样本名字list

    #生成文件路径
    path_class_counts = 'Data/object_class_counts'+str(times)+'.txt' #存放样本中已有的类别及统计的目标数目
    path_object_classes = 'Data/object_classes'+str(times)+'.txt'    #存放目标个数大于num的类别
    path_object_classes_to_value = 'Data/object_classes_to_value'+str(times)+'.txt'    #存放目标个数大于num的类别
    path_wrong_sample = 'Data/wrong_sample'+str(times)+'.txt'        #存放标注错误的样本路径

    wrong_sample = open(path_wrong_sample,'w')
    
    obejct_classes_counts = class_count(path_all_list , class_names)
    
    wrong_sample.close()

    f_object_class_counts = open(path_class_counts,'w')
    f_object_classes = open(path_object_classes,'w')
    f_object_classes_to_value = open(path_object_classes_to_value,'w')

    # 统计所有类别样本已有样本目标框数量
    object_all_count = open('Data/object_all_count'+times+'.txt','w')
    for i in class_names:
        if i in obejct_classes_counts:
            object_all_count.write('{} {} \n'.format(i, obejct_classes_counts[i]))
        else:
            object_all_count.write('{} {} \n'.format(i,	 0))

    object_all_count.close()

    i = 0
    cls_to_train = []
    for cls in class_names:
        if cls in obejct_classes_counts:
            if obejct_classes_counts[cls] >= num and cls not in class_not_train:
                cls_to_train.append(cls)
                log_file.write('****加入训练类别{:2d}:{:20s} \n'.format(i, cls))
                f_object_classes.write(cls+'\n')              #将大于num的目标数类别写入
                f_object_classes_to_value.write(cls+' '+class_to_value[cls]+'\n')  #将大于num的目标数类别及真值表写入
                i += 1
        else:
            continue


    #将数据写入xlsx
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = '训练数据日志'+times2
    
    cols=3
    rows=1
    
    # excel 写入class_name, 根据加入训练的数据顺序写
    sheet.cell(row=rows, column=2, value='class_name')
    for cls in cls_to_train:
        sheet.cell(row=rows, column=cols, value=cls)
        cols=cols+1
        
    for k in class_names:
        if k not in cls_to_train:
            sheet.cell(row=rows, column=cols, value=k)
            cols=cols+1
        else:
            continue
        
    # excel 写入values值
    value = []
    for i in cls_to_train:
        if i in class_to_value:
            value.append(class_to_value[i])
        else:
            continue
    
    for k in class_names:
        if k not in cls_to_train:
            value.append(class_to_value[i])
        else:
            continue
            
    rows=rows+1
    cols = 3
    sheet.cell(row=rows, column=2, value='values')
    for i in range(len(class_names)):
        sheet.cell(row=rows, column=cols, value=value[i])
        cols = cols + 1


    # excel写入所有类别对应的目标框个数
    value = []
    
    for i in cls_to_train:
        if i in obejct_classes_counts:
            value.append(obejct_classes_counts[i])
        else:
            value.append(0)    
    
    for i in class_names:
        if i not in cls_to_train:
            if i in obejct_classes_counts:
                value.append(obejct_classes_counts[i])
            else:
                value.append(0)
        else:
            continue

    rows= 3
    cols = 3
    sheet.cell(row=rows, column=2, value='目标框个数')
    for i in range(len(class_names)):
        sheet.cell(row=rows, column=cols, value=value[i])
        cols = cols + 1


    # excel写入类别是否加入训练
    value = []
    
    for i in cls_to_train:
        if i in obejct_classes_counts:
            if obejct_classes_counts[i] >= num and i not in class_not_train:
                value.append('Y')            
            else:
                value.append('N')            
        else:
            value.append('N')  
    
    for i in class_names:
        if i not in cls_to_train:
            if i in  obejct_classes_counts:
                if obejct_classes_counts[i] >= num and i not in class_not_train:
                    value.append('Y')
                else:
                    value.append('N')
            else:
                value.append('N')
        else:
            continue
    
    rows= 4
    cols = 3
    sheet.cell(row=rows, column=2, value='是否加入训练')
    for i in range(len(class_names)):
        sheet.cell(row=rows, column=cols, value=value[i])
        cols = cols + 1
    #
    obejct_classes_counts2 = sorted(obejct_classes_counts.items(), key=lambda x: (-x[1], x[0])) #将字典根据value值降序排列
    j = 0
    count = 0
    for key , value in obejct_classes_counts2:
        f_object_class_counts.write(key+' '+str(value)+'\n')  #写入样本中所有的类别及目标个数
        log_file.write('样本中类别{:2d}及个数:{:20s} {:6d} \n'.format(j, key , value))
        count += value
        j += 1

    f_object_class_counts.close()
    f_object_classes.close()
    f_object_classes_to_value.close()

    print('共有%s个目标框数目'%count)
    print('已有样本中类别及其目标个数: {}'.format(obejct_classes_counts))
    log_file.write('共有 %s 个目标框数目 \n'%count)
    log_file.write('已有样本中类别及个数: {} \n'.format(obejct_classes_counts))

    #计算train/val/test中每个类别目标框个数
    object_train_counts = object_class_count('Data/train_list.txt',class_names)
    object_val_counts = object_class_count('Data/val_list.txt',class_names)
    object_test_counts = object_class_count('Data/test_list.txt',class_names)
    log_file.write('训练集train各类别目标框个数: {} \n'.format(object_train_counts))
    log_file.write('验证集val  各类别目标框个数: {} \n'.format(object_val_counts))
    log_file.write('测试集test 各类别目标框个数: {} \n'.format(object_test_counts))
    print('训练集train各类别目标框个数: {} \n'.format(object_train_counts))
    print('验证集val  各类别目标框个数: {} \n'.format(object_val_counts))
    print('测试集test 各类别目标框个数: {} \n'.format(object_test_counts))

    # excel写入train/对应的各个目标框个数
    value = []
    for i in cls_to_train:
        if i in object_train_counts:
            value.append(object_train_counts[i]) 
        else:
            value.append(0)
    
    for i in class_names:
        if i not in cls_to_train:
            if i in object_train_counts:
                value.append(object_train_counts[i])
            else:
                value.append(0)
        else:
            continue
    
    rows= 5
    cols = 3
    sheet.cell(row=rows, column=2, value='train目标框')
    for i in range(len(class_names)):
        sheet.cell(row=rows, column=cols, value=value[i])
        cols = cols + 1
        
    # excel写入/val对应的各个目标框个数
    value = []
    
    for i in cls_to_train:
        if i in object_val_counts:
            value.append(object_val_counts[i]) 
        else:
            value.append(0)
    
    for i in class_names:
        if i not in cls_to_train:
            if i in object_val_counts:
                value.append(object_val_counts[i])          
            else:
                value.append(0)
        else:
            continue    
    
    rows= 6
    cols = 3
    sheet.cell(row=rows, column=2, value='val目标框')
    for i in range(len(class_names)):
        sheet.cell(row=rows, column=cols, value=value[i])
        cols = cols + 1

    # excel写入/test对应的各个目标框个数

    value = []
    for i in cls_to_train:
        if i in object_test_counts:
            value.append(object_test_counts[i]) 
        else:
            value.append(0)
    
    for i in class_names:
        if i not in cls_to_train:
            if i in object_test_counts:
                value.append(object_test_counts[i])  
            else:
                value.append(0)
        else:
            continue
    
    rows= 7
    cols = 3
    sheet.cell(row=rows, column=2, value='test目标框')
    for i in range(len(class_names)):
        sheet.cell(row=rows, column=cols, value=value[i])
        cols = cols + 1

    xlsx_name = '训练数据日志' + times2 + '.xlsx'
    # wb.save(os.path.join('D:\code\YOLOv3_keras\Data',xlsx_name))
    wb.save(os.path.join('Data',xlsx_name))
    print("写入数据成功！")
    
    log_file.write('本次样本中类别存放路径object_class_count: {} \n'.format(path_class_counts))
    log_file.write('本次选择目标数大于: {} 的类别放入训练 \n'.format(num))
    log_file.write('本次用于训练类别存放路径object_class: {} \n'.format(path_object_classes))
    log_file.write('本次用于训练类别真值表存放路径object_class_to_value: {} \n'.format(path_object_classes_to_value))
    print('本次用于训练类别存放路径object_class: {} \n'.format(path_object_classes))
    print('本次用于训练类别真值表存放路径object_class_to_value: {} \n'.format(path_object_classes_to_value))

def cal_class_size():
    id_class = {}
    classes = open('Data/object_classes'+str(times)+'.txt').read().strip().split()
    for i in classes:
        id_class[classes.index(i)] = i
    object_sizes2 = open('Data/object_sizes'+times+'.txt','w') #存放每个目标框尺寸
    object_sizes2.write('class_name, width, heigth, area \n')
    log_file.write('***各类别的目标框平均尺寸:***\n class_name, width, height, area \n')
    class_size = {}
    with open('Data/all_train_data.txt','r') as f:
        for items in f.readlines():
            for item in items.strip().split()[1:] :
                ids = int(item.split(',')[4])
                ids = id_class[ids]
                if ids not in class_size :
                    width = int(item.split(',')[2]) - int(item.split(',')[0])
                    height = int(item.split(',')[3]) - int(item.split(',')[1])
                    class_size[ids] = {'count':1 , 'W': width , 'H': height}
                else:
                    width = int(item.split(',')[2]) - int(item.split(',')[0])
                    height = int(item.split(',')[3]) - int(item.split(',')[1])
                    class_size[ids]['count'] += 1
                    class_size[ids]['W'] += width
                    class_size[ids]['H'] += height
    for key in class_size.keys():
        value = class_size[key]
        object_sizes2.write('{:20s} {:.1f} {:.1f} {:.1f}\n'.format(key, value['W']/value['count'], 
                                                                           value['H']/value['count'], value['W']*value['H']/value['count']/value['count']))
        log_file.write('{:20s} {:.1f} {:.1f} {:.1f}\n'.format(key, value['W']/value['count'], 
                                                                      value['H']/value['count'], value['W']*value['H']/value['count']/value['count']))
    object_sizes2.close()	

if __name__ == '__main__':

    s = \
        '#说明\n' +\
                '# all_list.txt : Data文件下所有样本的名字路径,不加.jpg和.xml\n'+\
                '# all_train_data.txt  : Data文件下所有图片的路径和图片的目标框对应的(xmin,ymin,xmax,ymax,label)\n'+\
                '# logs_file2018_10_20.txt : 程序运行日志信息\n'+\
                '# object_class_counts2018_10_20.txt : 样本中所有不同类别目标框个数统计\n'+\
                '# object_classes2018_10_20.txt : 本次统计目标框大于num的类别,放入训练模型\n'+\
                '# object_classes_to_value2018_10_20.txt :  本次统计目标框大于num的类别及真值表\n'+\
                '# object_sizes2018_10_20.txt : 根据all_train_data.txt文件, 统计放入训练的各类别的目标框平均尺寸\n'+\
                '# object_all_count2018_10_20.txt : 已有样本中统计47(包括shear2he/shear2fen)类目标对应的目标框个数\n'+\
                '# train_list.txt/ test_list.txt/ trainval_list.txt/ val_list.txt : 根据Data路径下每个不同文件夹中样本划分的样本集合\n'+\
                '# train.txt/ test.txt/ val.txt : 根据train_list.txt/ test_list.txt/ val_list.txt生成的训练测试数据\n'+\
                '# wrong_sample2018_10_20.txt : 存放标签标错的样本路径\n'

    start_time = time.time()
    localTime = time.localtime(time.time()) 
    times = time.strftime("%Y_%m_%d", localTime) 
    times2 = time.strftime("%Y_%m_%d_%H_%M_%S", localTime)
    print('本次运行程序时间:{} \n'.format(time.strftime("%Y-%m-%d %H:%M:%S", localTime)))	
    log_file = open('Data/logs_file'+times2+'.txt','a+')   #创建日志文档
    log_file.write('本次运行程序时间:{} \n'.format(time.strftime("%Y-%m-%d %H:%M:%S", localTime)))
    log_file.write(s)

    class_not_train = open('object_classes_not_train.txt','r').read().strip().split()  #不加入训练的类别标签
    log_file.write('不加入训练的类别:{} \n'.format(class_not_train))
    #删除Data路径下 train_list.txt /trainval_list.txt /val_list.txt /test_list.txt这4个.txt文件
    if os.path.exists('Data/train_list.txt'):
        os.remove('Data/train_list.txt')
    if os.path.exists('Data/trainval_list.txt'):
        os.remove('Data/trainval_list.txt')	
    if os.path.exists('Data/test_list.txt'):
        os.remove('Data/test_list.txt')	
    if os.path.exists('Data/val_list.txt'):
        os.remove('Data/val_list.txt')


    

    # 生成训练数据的函数
    creat_train_data_list()
    output_class_list(num = 100 , class_not_train = class_not_train)
    creat_train_data()
    cal_class_size()

    log_file.write('\n'+'='*1000 + '\n')
    log_file.write(s)

    log_file.close()

    
    
    print('用时(s):%s'%(time.time()-start_time))    