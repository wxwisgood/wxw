#!/usr/bin/env python
# -*- coding: gbk -*
# -*- coding: utf-8 -*-
# @Time    : 2018/10
# @Author  : WXW
# @Site    :
# @File    : .py
# @Software: PyCharm



import colorsys
import os
from timeit import default_timer as timer
import datetime
import numpy as np
from keras import backend as K
from keras.models import load_model
from keras.layers import Input
from PIL import Image, ImageFont, ImageDraw

from yolo3.model import yolo_eval, yolo_body, tiny_yolo_body
from yolo3.utils import letterbox_image
import os

from keras.utils import multi_gpu_model
# os.environ["CUDA_VISIBLE_DEVICES"] = "-1"

from pascal_map import com_mAP1,write_txt,ValidateFormats,ValidatePaths,getBoundingBoxes
import random
from  YOLO_tiny import YOLO
import openpyxl
# currentPath1 = 'E:\WXW\python\Python3_OpenCV3\work\YOLOv3_keras'
currentPath1= os.path.dirname(os.path.realpath(__file__))
model_paths=['logs/20181020/yolov3_tiny_20181020_1.h5','logs/20181020/yolov3_tiny_20181020_2.h5',
            'logs/20181020/yolov3_tiny_20181020_3.h5','logs/20181020/yolov3_tiny_20181020_4.h5',
            'logs/20181020/yolov3_tiny_20181020_5.h5','logs/20181020/yolov3_tiny_20181020_6.h5',
              'logs/20181020/yolov3_tiny_20181020_7.h5']
scores=[0.2,0.3,0.4]
ious=[0.2]
sizes=[288,416,608]

if __name__=='__main__':

    #准备所需目标类别，文件列表
    object_txt = os.path.join(currentPath1, 'Data\\test.txt')
    classes_path=os.path.join(currentPath1, 'Data/object_classes2018_10_20.txt')

    classes_recongition_value_path=os.path.join(currentPath1, "Data/object_classes_to_value2018_10_20.txt")
    anchors_path= os.path.join(currentPath1, 'model_data/yolov3_tiny_anchors_416_20181020.txt')


    with open(classes_path) as f:
        class_names = f.readlines()
    for i in range(len(class_names)):
        class_names[i] = class_names[i][:-1]

    with open(object_txt) as f:
        lines = f.readlines()
    # 将数据写入xlsx
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'mAP测试结果'

    cols = 4
    rows = 1



    dict_mAP = dict()
    dict_mAP['model_path'] = 0
    dict_mAP['Image_size'] = 0
    dict_mAP['iou'] = 0
    dict_mAP['score'] = 0
    dict_mAP['mAP'] = 0
    for class_name in class_names:
        dict_mAP[class_name] = 0
    for k in dict_mAP.keys():
        sheet.cell(row=rows, column=cols, value=k)
        cols = cols + 1
    rows = rows + 1
    time_str = datetime.datetime.now().strftime('%Y%m%d-%H%M%S')
    xlsx_name = 'mAP/mAP测试结果' + time_str + '.xlsx'
    wb.save(os.path.join(currentPath1, xlsx_name))

    dict_mAP_all = []
    for model_path in model_paths:
        ave_map = []
        for size in sizes:
            for iou in ious:
                for score in scores:
                    K.clear_session()
                    # 加载YOLO模型
                    model_path=os.path.join(currentPath1,model_path)
                    yolo=YOLO(model_path=model_path,score=score,iou=iou,classes_path=classes_path,
                              classes_recongition_value_path=classes_recongition_value_path,anchors_path=anchors_path,model_image_size=(size,size))
                    model_name = model_path.split('/')[-1]

                    print("Now Process model:{} iou:{}  score:{} size:{}".format(model_name,iou,score,(size,size)))

                    for line in lines:
                        line = line[:-1]
                        a = line.split(' ')
                        image_name = a[0]

                        image_name = os.path.join(currentPath1,image_name)
                        image = Image.open(image_name)
                        out_boxes, out_scores, out_classes = yolo.detect_image_txt(image)
                        image_name1=(image_name.split('\\')[-1]).split('.')[0]
                        with open((currentPath1+'/mAP/Object-Detection-Metrics-master/detections/'+image_name1+".txt"),'w+') as f1:
                            write_txt(out_boxes, out_scores, out_classes,f1,class_names)
                    dict_mAP=dict()
                    dict_mAP['model_path'] = model_name
                    dict_mAP['Image_size'] = size
                    dict_mAP['iou'] = iou
                    dict_mAP['score'] = score
                    dict_mAP['mAP'] = 0
                    for class_name in class_names:
                        dict_mAP[class_name]=0


                    dict_mAP = com_mAP1(dict_mAP)
                    print(dict_mAP)
                    dict_mAP_all.append(dict_mAP)

                    cols = 4
                    for k in dict_mAP.keys():
                        sheet.cell(row=rows, column=cols, value=dict_mAP[k])
                        cols = cols + 1
                    rows = rows + 1
                    wb.save(os.path.join(currentPath1, xlsx_name))
    global rows_ave
    rows_ave = 1
    def write_ave_comu(sheet, name, name_list, dict_mAP_all):
        global rows_ave
        sheet.cell(row=rows_ave, column=1, value=name)
        rows_ave += 1
        for model_path in name_list:
            if(name=='model_path'):
                model_path = model_path.split('/')[-1]
            sheet.cell(row=rows_ave, column=1, value=model_path)

            ave_sum = []
            for data in dict_mAP_all:
                if (data[name] == model_path):
                    ave_sum.append(float(data['mAP']))
                avesum = np.mean(np.array(ave_sum))
            sheet.cell(row=rows_ave, column=2, value=avesum)
            rows_ave += 1



    write_ave_comu(sheet, name='model_path', name_list=model_paths,dict_mAP_all=dict_mAP_all)
    write_ave_comu(sheet, name='Image_size', name_list=sizes,dict_mAP_all=dict_mAP_all)
    write_ave_comu(sheet, name='iou', name_list=ious,dict_mAP_all=dict_mAP_all)
    write_ave_comu(sheet, name='score', name_list=scores,dict_mAP_all=dict_mAP_all)

    wb.save(os.path.join(currentPath1, xlsx_name))
    print("写入数据成功！")