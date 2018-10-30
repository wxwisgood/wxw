#!/usr/bin/env python
# -*- coding: gbk -*
# -*- coding: utf-8 -*-
# @Time    : 2018/10
# @Author  : WXW
# @Site    :
# @File    : .py
# @Software: PyCharm


import colorsys
import os
from timeit import default_timer as timer
import datetime
import numpy as np
from keras import backend as K
from keras.models import load_model
from keras.layers import Input
from PIL import Image, ImageFont, ImageDraw

from yolo3.model import yolo_eval, yolo_body, tiny_yolo_body
from yolo3.utils import letterbox_image
import os

from keras.utils import multi_gpu_model
# os.environ["CUDA_VISIBLE_DEVICES"] = "-1"

from pascal_map import com_mAP1,write_txt,ValidateFormats,ValidatePaths,getBoundingBoxes
import random
from  YOLO import YOLO
import openpyxl
# currentPath1 = 'E:\WXW\python\Python3_OpenCV3\work\YOLOv3_keras'
currentPath1= os.path.dirname(os.path.realpath(__file__))
model_paths=['logs/20181024_1080Ti/yolov3_20181024_4.h5']
# model_paths=['logs/201810241/yolov3_20181024_taishi_1.h5','logs/201810241/yolov3_20181024_taishi_2.h5',
#              'logs/201810241/yolov3_20181024_taishi_3.h5']
scores=[0.05]
ious=[0.2]
# sizes=[416]
use_soft_nmses=[False]
# scores=[0.2]
# ious=[0.2]
sizes=[320]
# use_soft_nmses=[False,True]

if __name__=='__main__':

    #准备所需目标类别，文件列表
    object_txt = os.path.join(currentPath1, 'Data\\test.txt')
    classes_path = os.path.join(currentPath1, 'model_data/object_classes2018_10_24.txt')

    classes_recongition_value_path = os.path.join(currentPath1, "model_data/object_classes_to_value2018_10_24.txt")

    anchors_path= os.path.join(currentPath1, 'model_data/yolov3_anchors_416_20181020.txt')


    with open(classes_path) as f:
        class_names = f.readlines()
    for i in range(len(class_names)):
        class_names[i] = class_names[i][:-1]

    with open(object_txt) as f:
        lines = f.readlines()
    # 将数据写入xlsx
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'mAP测试结果'

    cols = 4
    rows = 2#从第二行开始写，第一行留给平均值
    dict_mAP = dict()
    dict_mAP['model_path'] = 0
    dict_mAP['Image_size'] = 0
    dict_mAP['iou'] = 0
    dict_mAP['score'] = 0
    dict_mAP['use_soft_nms'] = True
    dict_mAP['mAP'] = 0
    for class_name in class_names:
        dict_mAP[class_name] = 0
    for k in dict_mAP.keys():
        sheet.cell(row=rows, column=cols, value=k)
        cols = cols + 1
    rows = rows + 1
    time_str = datetime.datetime.now().strftime('%Y%m%d-%H%M%S')
    xlsx_name = 'mAP/mAP测试结果' + time_str + '.xlsx'
    wb.save(os.path.join(currentPath1, xlsx_name))

    dict_mAP_all = []
    for model_path in model_paths:
        ave_map = []
        for size in sizes:
            for iou in ious:
                for score in scores:
                    for use_soft_nms in use_soft_nmses:
                        K.clear_session()
                        # 加载YOLO模型
                        model_path=os.path.join(currentPath1,model_path)
                        yolo=YOLO(model_path=model_path,score=score,iou=iou,classes_path=classes_path,
                                  classes_recongition_value_path=classes_recongition_value_path,anchors_path=anchors_path,
                                  model_image_size=(size,size),Use_Soft_NMS=use_soft_nms)
                        model_name = model_path.split('/')[-1]

                        print("Now Process model:{} iou:{}  score:{} size:{} Use_Soft_NMS:{}".format(model_name,iou,score,(size,size),use_soft_nms))

                        for line in lines:
                            line = line[:-1]
                            a = line.split(' ')
                            image_name = a[0]

                            image_name = os.path.join(currentPath1,image_name)
                            image = Image.open(image_name)
                            out_boxes, out_scores, out_classes = yolo.detect_image_txt(image)
                            image_name1=(image_name.split('\\')[-1]).split('.')[0]
                            with open((currentPath1+'/mAP/Object-Detection-Metrics-master/detections/'+image_name1+".txt"),'w+') as f1:
                                write_txt(out_boxes, out_scores, out_classes,f1,class_names)
                        dict_mAP=dict()
                        dict_mAP['model_path'] = model_name
                        dict_mAP['Image_size'] = size
                        dict_mAP['iou'] = iou
                        dict_mAP['score'] = score
                        dict_mAP['use_soft_nms']=use_soft_nms
                        dict_mAP['mAP'] = 0
                        for class_name in class_names:
                            dict_mAP[class_name]=0


                        dict_mAP = com_mAP1(dict_mAP)
                        print(dict_mAP)
                        dict_mAP_all.append(dict_mAP)

                        cols = 4
                        for k in dict_mAP.keys():
                            sheet.cell(row=rows, column=cols, value=dict_mAP[k])
                            cols = cols + 1
                        rows = rows + 1
                        wb.save(os.path.join(currentPath1, xlsx_name))
    global rows_ave
    rows_ave = 2
    #计算每一个模型或者尺度的mAP平均值
    def write_ave_comu(sheet, name, name_list, dict_mAP_all):
        global rows_ave
        sheet.cell(row=rows_ave, column=1, value=name)
        rows_ave += 1
        for model_path in name_list:
            if(name=='model_path'):#model_path 特殊处理
                model_path = model_path.split('/')[-1]
            #写名称
            sheet.cell(row=rows_ave, column=1, value=model_path)

            #计算平均值
            ave_sum = []
            for data in dict_mAP_all:
                if (data[name] == model_path):
                    ave_sum.append(float(data['mAP']))
            avesum = np.mean(np.array(ave_sum))
            sheet.cell(row=rows_ave, column=2, value=avesum)
            rows_ave += 1

    write_ave_comu(sheet, name='model_path', name_list=model_paths,dict_mAP_all=dict_mAP_all)
    write_ave_comu(sheet, name='Image_size', name_list=sizes,dict_mAP_all=dict_mAP_all)
    write_ave_comu(sheet, name='iou', name_list=ious,dict_mAP_all=dict_mAP_all)
    write_ave_comu(sheet, name='score', name_list=scores,dict_mAP_all=dict_mAP_all)
    write_ave_comu(sheet, name='use_soft_nms', name_list=use_soft_nmses, dict_mAP_all=dict_mAP_all)

    # 计算每一类目标的平均mAP
    row=1 #写在第一行
    def write_ave_object_mAP_comu(name,row,col,dict_mAP_all):

        # 计算平均值
        ave_sum = []
        for data in dict_mAP_all:
            ave_sum.append(data[name])
        avesum = np.mean(np.array(ave_sum))
        sheet.cell(row=row, column=col, value=avesum)



    cols = 4
    # 不需要计算平均值的key
    set_no_ave = {'model_path', 'Image_size', 'iou', 'score','use_soft_nms'}
    for k in dict_mAP.keys():
        #判断是否需要计算
        if(k not in set_no_ave):
            write_ave_object_mAP_comu(name=k, row=row, col=cols, dict_mAP_all=dict_mAP_all)
        cols = cols + 1


    wb.save(os.path.join(currentPath1, xlsx_name))
    print("写入数据成功！")